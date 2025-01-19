import openpyxl

Dict = {}

book = openpyxl.load_workbook("C:/Users/nages/PycharmProjects/PythonSelfFramework/testdata/pythonDemo.xlsx")
sheet = book.active

book = openpyxl.load_workbook("C:/Users/nages/PycharmProjects/PythonSelfFramework/testdata/pythonDemo.xlsx")
sheet = book.active

all_row = []

for i in range(2, sheet.max_row + 1):
    Dict = {}
    for j in range(1, sheet.max_column + 1):
        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    all_row.append(Dict)
print(all_row)



