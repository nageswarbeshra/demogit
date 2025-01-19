import openpyxl
import pytest

from testdata.testCases import Testcases


class HomePageData:


    # test_home_page = [{"firstname":"rahul","lastname":"shetty","gender":"Male"},{"firstname":"Anshika","lastname":"shetty","gender":"Female"}]

    @staticmethod
    def getTestData():


        book = openpyxl.load_workbook("C:/Users/nages/PycharmProjects/PythonSelfFramework/testdata/test_data50.xlsx")
        sheet = book.active

        all_row = []

        for i in range(2, sheet.max_row + 1):
            Dict = {}
            for j in range(1, sheet.max_column + 1):
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            all_row.append(Dict)
        return all_row




