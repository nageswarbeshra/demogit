import openpyxl


class Testcases:

    def getTestcases(self):

        book = openpyxl.load_workbook("C:/Users/nages/PycharmProjects/PythonSelfFramework/testdata/test_data.xlsx")
        sheet = book.active

        for i in range(2, sheet.max_row + 1):
            for j in range(1, 2):
                testcasess = sheet.cell(row=i, column=j).value
                return testcasess




